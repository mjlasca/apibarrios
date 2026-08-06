#!/usr/bin/env python3
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font


def main():
    if len(sys.argv) != 3:
        print("Usage: generate_xlsx.py <input.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    json_path = sys.argv[1]
    xlsx_path = sys.argv[2]

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    headers = data.get('headers', [])
    rows = data.get('rows', [])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    bold_font = Font(bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(xlsx_path)


if __name__ == '__main__':
    main()
